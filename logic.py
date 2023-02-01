from blizzardapi import BlizzardApi
import datetime
import requests
import urllib.parse
import openpyxl
from openpyxl.styles import PatternFill, Font
import db

client_id = ''
client_secret = ''

class ApiLauncher:
    all_chars = 0

    def __init__(self):
        pass

    def get_tokens(self):
        api_client = BlizzardApi(client_id.strip(), client_secret.strip())
        try:
            test_for_access = api_client.wow.game_data.get_mounts_index('eu', 'ru_RU')
        except:
            return 'Ошибка'

    def get_char_info(self, char_name):
        api_client = BlizzardApi(client_id.strip(), client_secret.strip())
        char_info = api_client.wow.profile.get_character_profile_summary('eu', 'ru_RU', 'howling-fjord', char_name)
        return char_info

    def get_equip_api(self, char_name):
        api_client = BlizzardApi(client_id.strip(), client_secret.strip())
        equip = api_client.wow.profile.get_character_equipment_summary('eu', 'ru_RU', 'howling-fjord', char_name)
        return equip

    @staticmethod
    def get_mythic_api(char_name):
        all_mythic = 0
        name = urllib.parse.quote(char_name)
        response = requests.get(
            f'https://raider.io/api/v1/characters/profile?region=eu&realm=howling%20fjord&name={name}&fields=mythic_plus_recent_runs')
        json_response = response.json()
        now_time = datetime.datetime.today()
        today = now_time.isoweekday()
        found_number = today - 3
        now_time = datetime.datetime.timetuple(now_time)
        for item in json_response['mythic_plus_recent_runs']:
            dungeon_time = item['completed_at'].split('.')
            dungeon_time = dungeon_time[0].replace('T', '-')
            dungeon_time = datetime.datetime.strptime(dungeon_time, '%Y-%m-%d-%H:%M:%S')
            dungeon_time = datetime.datetime.timetuple(dungeon_time)
            if now_time[2] - dungeon_time[2] > found_number or dungeon_time[3] < 7 or now_time[1] != dungeon_time[1]:
                continue
            all_mythic += 1
        return all_mythic

    @staticmethod
    def get_item_level(equip):
        all_lvl = 0
        all_items = 0
        for counter, item in enumerate(equip['equipped_items']):
            try:
                if item['slot']['type'] == 'SHIRT' or item['slot']['type'] == 'TABARD':
                    continue
            except:
                pass
            item_level = item['level']['value']
            all_items += 1
            all_lvl += item_level
        mid_level = all_lvl / all_items
        return mid_level

    @staticmethod
    def get_set_items(equip, char):
        all_sets_items = 0
        head, shoulder, chest, legs, hands = '-', '-', '-', '-', '-'
        for counter, item in enumerate(equip['equipped_items']):
            try:
                if len(item['set']['items']) == 5:
                    all_sets_items += 1
                    item_type = item['slot']['type']
                    item_lvl = item['level']['value']
                    if item_type == 'HEAD':
                        head= item_lvl
                    if item_type == 'SHOULDER':
                        shoulder = item_lvl
                    if item_type == 'CHEST':
                        chest = item_lvl
                    if item_type == 'LEGS':
                        legs = item_lvl
                    if item_type == 'HANDS':
                        hands = item_lvl
            except:
                pass
        char.char_set_head = head
        char.char_set_shoulder = shoulder
        char.char_set_chest = chest
        char.char_set_legs = legs
        char.char_set_hands = hands
        char.char_set_count = f'{all_sets_items}/5'
        char.save()

    def get_info(self, g_static):
        items = g_static.select()
        for char in items:
            char_name = char.char_name
            try:
                char_equip = self.get_equip_api(char_name.lower())
                char_class = self.get_char_info(char_name.lower())['character_class']['name']
                ilvl = self.get_item_level(char_equip)
            except:
                return 'Ошибка'
            ilvl = float('{:.2f}'.format(ilvl))
            mythic_done = self.get_mythic_api(char_name)
            char.char_itemlvl = ilvl
            char.char_mythic_done = mythic_done
            char.char_class = char_class
            char.save()
            self.get_set_items(char_equip, char)
            self.all_chars += 1
        self.all_chars = 0
        return 'Обновление выполнено'


def full_update():
    launcher = ApiLauncher()
    first = launcher.get_info(db.Static)
    if first == 'Обновление выполнено':
        return first
    else:
        return 'ОШИБКА'


def add_player(name, role):
    db.Static.create(char_name=name, char_role=role)
    launcher = ApiLauncher()
    response = launcher.get_info(db.Static.select().order_by(db.Static.char_id.desc()).get())
    return response


def table_style():
    class_colors = {'Воин': 'C69B6D',
                    'Паладин': 'F48CBA',
                    'Охотник': 'AAD372',
                    'Разбойник': 'FFF468',
                    'Жрец': 'FFFFFF',
                    'Шаман': '0070DD',
                    'Маг': '3FC7EB',
                    'Чернокнижник': '8788EE',
                    'Монах': '00FF98',
                    'Друид': 'FF7C0A ',
                    'Охотник на демонов': 'A330C9',
                    'Рыцарь смерти': 'C41E3A ',
                    'Пробудитель': '33937F '}
    book = openpyxl.load_workbook('Статик.xlsx')
    sheet = book.active

    #Размер ячеек
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["L"].width = 18

    #Заливка цветом класса
    for item in sheet['D']:
        if item.value in class_colors:
            item.font = Font(b=True, outline=True, color='000000')
            item.fill = PatternFill("solid", fgColor=class_colors[item.value])
            print(item.value)

    #Изменение цвета текста в зависимости от уровня предмета
    def quality_colors(sheet_cells):
        for item in sheet_cells:
            try:
                ilvl = float(item.value)
                if 390 <= ilvl < 405:
                    item.font = Font(color='a335ee')
                elif 380 <= ilvl < 390:
                    item.font = Font(color='0070dd')
                elif ilvl > 405:
                    item.font = Font(color='ff8000')
            except:
                pass

    quality_colors(sheet['E'])
    quality_colors(sheet['G'])
    quality_colors(sheet['H'])
    quality_colors(sheet['I'])
    quality_colors(sheet['J'])
    quality_colors(sheet['K'])

    book.save('Статик.xlsx')
